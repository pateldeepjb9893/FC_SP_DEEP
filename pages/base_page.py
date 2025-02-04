from selenium.webdriver import ActionChains
from locators.vane_county_locators import VaneCountyLocator
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from difflib import SequenceMatcher
import os, pdfplumber, time, requests, pytesseract, pandas as pd, urllib.parse


class BasePage:
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)

    def click(self, locator):
        element = self.driver.find_element(*locator)
        ActionChains(self.driver).move_to_element(element).perform()
        self.wait.until(EC.element_to_be_clickable(locator)).click()

    def send_keys(self, locator, text):
        self.wait.until(EC.visibility_of_element_located(locator)).send_keys(text)

    def get_text(self, locator):
        return self.wait.until(EC.visibility_of_element_located(locator)).text

    def switch_to_frame(self, frame_locator):
        iframe = self.wait.until(EC.presence_of_element_located(frame_locator))
        self.driver.switch_to.frame(iframe)
        print(f"Switched to iframe: {frame_locator}")

    def select_dropdown_option(self, dropdown_locator, option_value, by=By.ID, select_type="visible_text"):
        dropdown_element = self.driver.find_element(by, dropdown_locator)
        dropdown = Select(dropdown_element)

        if select_type == "visible_text":
            dropdown.select_by_visible_text(option_value)
        elif select_type == "index":
            dropdown.select_by_index(option_value)
        elif select_type == "value":
            dropdown.select_by_value(option_value)
        else:
            raise ValueError("Invalid select_type. Choose 'visible_text', 'index', or 'value'.")

    def send_date(self, locator):
        today_date = (datetime.now() - timedelta(days=abs(19))).strftime("%m/%d/%Y")
        self.wait.until(EC.visibility_of_element_located(locator))
        # Clear the existing value before sending the new date
        input_field = self.driver.find_element(*locator)
        input_field.clear()
        action = ActionChains(self.driver)
        action.move_to_element(input_field).click().send_keys(today_date).perform()
        print(f"Sent date: {today_date} to element: {locator}")

    def take_screenshot(self, screenshot_name):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        screenshot_directory = "screenshots"

        if not os.path.exists(screenshot_directory):
            os.makedirs(screenshot_directory)
        screenshot_path = os.path.join(screenshot_directory, f"{screenshot_name}_{timestamp}.png")

        self.driver.save_screenshot(screenshot_path)
        print(f"Screenshot saved at: {screenshot_path}")

        return screenshot_path

    def switch_to_new_window(self):
        """Switch to the newly opened browser window."""
        main_window = self.driver.current_window_handle
        for handle in self.driver.window_handles:
            if handle != main_window:
                self.driver.switch_to.window(handle)
                break

    def extract_pdf_text_from_new_window(self):
        try:
            self.switch_to_new_window()
            time.sleep(5)
            pdf_url = self.driver.current_url
            response = requests.get(pdf_url)
            if response.status_code != 200:
                raise Exception(f"Failed to download PDF. Status code: {response.status_code}")
            pdf_path = "temp_downloaded_file.pdf"
            with open(pdf_path, "wb") as pdf_file:
                pdf_file.write(response.content)
            extracted_lines = []
            pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
            output_dir = "ocr_text"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            output_text_path = os.path.join(output_dir, f"ocr_output_{timestamp}.txt")
            with open(output_text_path, "w", encoding="utf-8") as output_file:
                with pdfplumber.open(pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        img = page.to_image(resolution=300).original
                        text = pytesseract.image_to_string(img)
                        if text:
                            page_header = f"--- Page {page_num + 1} ---\n"
                            output_file.write(page_header)
                            output_file.write(text)
                            output_file.write("\n\n")
                            extracted_lines.append(page_header.strip())
                            extracted_lines.extend(text.splitlines())
            os.remove(pdf_path)
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
            return extracted_lines
        except Exception as e:
            print(f"Error processing page")

    @staticmethod
    def extract_address_from_ocr(file_path):
        address_found = False
        address_lines = []

        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()

                if line.startswith("Commonly known as:"):
                    return line.replace("Commonly known as:", "").strip()

                if address_found:
                    if line:
                        address_lines.append(line)
                        if len(address_lines) == 2:
                            break

                if "The common address or location of the property is" in line:
                    address_found = True

        return " ".join(address_lines) if address_lines else "Not Found"

    def process_ocr_files_and_save_to_excel(self, directory="ocr_text"):
        data = []
        for filename in os.listdir(directory):
            if filename.endswith(".txt"):
                file_path = os.path.join(directory, filename)
                address = self.extract_address_from_ocr(file_path)
                data.append({"File Name": filename, "Extracted Address": address})
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_directory = "propertyDetails"
        if not os.path.exists(file_directory):
            os.makedirs(file_directory)
        file_name_time = f"propertyDetails_{timestamp}.xlsx"
        excel_filename = os.path.join(file_directory, file_name_time)
        df.to_excel(excel_filename, index=False)
        print(f"Excel file '{excel_filename}' created successfully!")
        self.generate_search_url_from_excel(excel_filename)

    @staticmethod
    def generate_search_url_from_excel(excel_file):
        df = pd.read_excel(excel_file)

        if "Extracted Address" not in df.columns:
            raise ValueError("Excel file must contain a column named 'Extracted Address'")

        base_url = "https://www.truepeoplesearch.com/resultaddress?"
        search_urls = []

        for address in df["Extracted Address"].dropna():
            try:
                parts = address.split(", ")
                street_address = parts[0]
                city_state_zip = ", ".join(parts[1:])

                street_encoded = urllib.parse.quote(street_address)
                city_state_zip_encoded = urllib.parse.quote(city_state_zip)

                search_url = f"{base_url}streetaddress={street_encoded}&citystatezip={city_state_zip_encoded}"
                search_urls.append(search_url)

            except Exception as e:
                print(f"Error processing address: {address} - {e}")
                search_urls.append("Error generating URL")

        df["Search URL"] = search_urls

        with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")

        print(f"Updated Excel file with search URLs: {excel_file}")

    def scroll_to_element(self, locator):
        element = self.driver.find_element(*locator)
        self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        ActionChains(self.driver).move_to_element(element).perform()
        print(f"Scrolled to element: {locator}")

    def extract_and_store_names(self, excel_file):
        try:
            all_name_elements = self.driver.find_elements(*VaneCountyLocator.NAME_LOCATOR)
            all_names = [element.text.strip() for element in all_name_elements if element.text.strip()]

            if not all_names:
                print("No names found.")
                return

            print(f"Extracted Names: {all_names}")

            book = load_workbook(excel_file)
            sheet = book.active
            df = pd.read_excel(excel_file)

            if "Defendant Names" not in df.columns:
                df["Defendant Names"] = ""

            df.at[0, "Defendant Names"] = ", ".join(all_names)

            with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")

            print(f"Stored names in Excel: {excel_file}")

        except Exception as e:
            print(f"Error in extract_and_store_names method: {e}")

    @staticmethod
    def flatten_and_convert(value):
        if isinstance(value, list):
            if all(isinstance(i, list) for i in value):  # If it's a list of lists, flatten it
                value = [item for sublist in value for item in sublist]
            return ", ".join(map(str, value))  # Convert list to comma-separated string
        return value  # Return as is if not a list

    def create_or_update_excel(self, data_map, folder_path="propertyDetails"):
        try:
            current_date = datetime.now().strftime("%Y-%m-%d")
            file_name = f"PropertyDetails_{current_date}.xlsx"
            file_path = os.path.join(folder_path, file_name)

            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            for key, value in data_map.items():
                data_map[key] = self.flatten_and_convert(value)

            if os.path.exists(file_path):
                print(f"File already exists. Updating file: {file_path}")
                wb = load_workbook(file_path)
                ws = wb.active

                df = pd.DataFrame([list(data_map.values())], columns=data_map.keys())
                ws.append(list(data_map.values()))

                wb.save(file_path)

            else:
                print(f"Creating new file: {file_path}")
                wb = Workbook()
                ws = wb.active

                ws.append(list(data_map.keys()))
                ws.append(list(data_map.values()))
                wb.save(file_path)

            print(f"Data successfully written to: {file_path}")

        except Exception as e:
            print(f"Error creating or updating Excel file: {e}")

    # def search_and_store_phone_numbers(self, excel_file):
    #     try:
    #         book = load_workbook(excel_file)
    #         sheet = book.active
    #         df = pd.read_excel(excel_file)
    #
    #         if "Defendant Names" not in df.columns or df["Defendant Names"].isnull().all():
    #             print("No names found in Excel.")
    #             return
    #
    #         if "Phone Number" not in df.columns:
    #             df["Phone Number"] = ""
    #
    #         all_names = df.at[0, "Defendant Names"].split(", ")
    #         print(f"Names from Excel: {all_names}")
    #
    #         extracted_addresses = df["Extracted Address"].dropna().tolist()
    #         print(f"Extracted Addresses from Excel: {extracted_addresses}")
    #
    #         phone_numbers = []
    #
    #         for name in all_names:
    #             base_url = "https://www.truepeoplesearch.com/results?"
    #             encoded_name = urllib.parse.quote(name)
    #             search_url = f"{base_url}name={encoded_name}"
    #
    #             print(f"Opening URL: {search_url}")
    #             self.driver.get(search_url)
    #             time.sleep(5)
    #
    #             try:
    #                 address_element = self.driver.find_element(By.XPATH,
    #                                                            "//div[contains(@class, 'card-content')]//a[contains(@href, 'resultaddress')]")
    #                 found_address = address_element.text.strip()
    #                 print(f"Found Address: {found_address}")
    #
    #                 # Find best matching address from Excel
    #                 best_match = None
    #                 best_similarity = 0
    #
    #                 for excel_address in extracted_addresses:
    #                     similarity = SequenceMatcher(None, excel_address, found_address).ratio()
    #                     print(f"Comparing with: {excel_address} | Similarity: {similarity}")
    #
    #                     if similarity > best_similarity:
    #                         best_similarity = similarity
    #                         best_match = excel_address
    #
    #                 if best_similarity >= 0.6:  # Acceptable similarity threshold
    #                     print(f"Best matching address: {best_match} | Similarity: {best_similarity}")
    #
    #                     # Step 4: Extract phone number
    #                     try:
    #                         phone_element = self.driver.find_element(By.XPATH, "//div[contains(@class, 'phone')]//a")
    #                         phone_number = phone_element.text.strip()
    #                     except:
    #                         phone_number = "Not Found"
    #
    #                     phone_numbers.append(phone_number)
    #                     print(f"Phone Number: {phone_number}")
    #
    #                 else:
    #                     print("No matching address found. Skipping phone number extraction.")
    #                     phone_numbers.append("Not Found")
    #
    #             except Exception as e:
    #                 print(f"Error extracting address or phone number: {e}")
    #                 phone_numbers.append("Not Found")
    #
    #         # Step 5: Store phone numbers in Excel (comma-separated in one cell)
    #         df.at[0, "Phone Number"] = ", ".join(phone_numbers)
    #
    #         with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    #             df.to_excel(writer, index=False, sheet_name="Sheet1")
    #
    #         print(f"Updated Excel file with phone numbers: {excel_file}")
    #
    #     except Exception as e:
    #         print(f"Error in search_and_store_phone_numbers method: {e}")
